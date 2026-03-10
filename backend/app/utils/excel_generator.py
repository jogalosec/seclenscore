"""
Generador de archivos Excel con openpyxl.
Equivalente a los usos de PhpSpreadsheet en index.php.
"""
import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _estilo_cabecera(ws, fila: int = 1) -> None:
    """Aplica estilo de cabecera (negrita + fondo azul) a la primera fila."""
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[fila]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = Alignment(horizontal="center")


def generar_excel_arbol_activos(arbol: List[Any]) -> bytes:
    """
    Genera un Excel con el árbol de activos.
    Equivalente a downloadActivosTree en index.php.

    El primer elemento de `arbol` son las cabeceras (lista de strings).
    Los siguientes son dicts con nombre, tipo, id, padre, archivado, expuesto.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Árbol de Activos"

    for fila_idx, fila in enumerate(arbol, start=1):
        if isinstance(fila, list):
            # Cabeceras
            ws.append(fila)
            _estilo_cabecera(ws, fila_idx)
        elif isinstance(fila, dict):
            ws.append([
                fila.get("nombre", ""),
                fila.get("tipo", ""),
                fila.get("id", ""),
                fila.get("padre", ""),
                "Sí" if fila.get("archivado") else "No",
                "Sí" if fila.get("expuesto") else "No",
            ])

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_excel_activos(activos: List[Dict], nombre_hoja: str = "Activos") -> bytes:
    """
    Genera un Excel con lista plana de activos.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    cabeceras = ["ID", "Nombre", "Tipo", "Usuario", "Descripción", "Archivado", "Expuesto", "Crítico"]
    ws.append(cabeceras)
    _estilo_cabecera(ws)

    for activo in activos:
        ws.append([
            activo.get("id"),
            activo.get("nombre"),
            activo.get("tipo_nombre") or activo.get("tipo"),
            activo.get("user_id"),
            activo.get("descripcion", ""),
            "Sí" if activo.get("archivado") else "No",
            "Sí" if activo.get("expuesto") else "No",
            "Sí" if activo.get("critico") else "No",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_excel_generico(
    datos: List[Dict],
    cabeceras: List[str],
    nombre_hoja: str = "Datos",
    campos: Optional[List[str]] = None,
) -> bytes:
    """
    Genera un Excel genérico a partir de una lista de dicts.
    Reutilizable para cualquier módulo que necesite exportar datos.
    """
    if campos is None:
        campos = list(datos[0].keys()) if datos else []

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]  # Excel limita a 31 caracteres

    ws.append(cabeceras)
    _estilo_cabecera(ws)

    for fila in datos:
        ws.append([fila.get(campo, "") for campo in campos])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
